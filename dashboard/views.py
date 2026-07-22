from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView
from django.db.models import Count, Q, Sum
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.decorators.http import require_POST

from accounts.models import PasswordResetRequest
from catalog.models import Banner, Category, Product
from core.models import ActivityLog, Coupon, PaymentMethod, ShippingZone, SiteConfiguration, TaxRate
from core.utils import log_activity
from orders.models import Order

from .forms import (
    BannerForm,
    CategoryForm,
    CouponForm,
    DashboardAuthenticationForm,
    PaymentMethodForm,
    ProductForm,
    ShippingZoneForm,
    SiteConfigurationForm,
    TaxRateForm,
    UserDashboardForm,
)


def _staff_check(user):
    return user.is_authenticated and user.is_staff


staff_required = user_passes_test(_staff_check, login_url=reverse_lazy("dashboard:login"))


class DashboardLoginView(LoginView):
    """
    Connexion au dashboard.
    Après authentification réussie, enregistre l'ID utilisateur
    dans la session sous `_dashboard_authenticated_user_id`.
    Ceci permet au DashboardAuthMiddleware de vérifier que l'utilisateur
    est bien autorisé à accéder au dashboard.
    """
    form_class = DashboardAuthenticationForm
    template_name = "dashboard/login.html"
    redirect_authenticated_user = True

    def form_valid(self, form):
        """Connecte l'utilisateur ET marque la session pour le dashboard."""
        response = super().form_valid(form)
        # Logger la connexion dashboard
        log_activity(
            self.request.user, "login", "Dashboard",
            None, self.request.user.username,
            f"Connexion au tableau de bord",
            self.request.META.get("REMOTE_ADDR"),
        )
        # Marquer la session comme authentifiée pour le dashboard
        self.request.session["_dashboard_authenticated_user_id"] = self.request.user.id
        self.request.session.save()
        return response

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.is_staff:
            # Vérifier si l'indicateur dashboard est déjà dans la session
            dashboard_uid = request.session.get("_dashboard_authenticated_user_id")
            if dashboard_uid == request.user.id:
                return redirect("dashboard:index")
            # Sinon, on le définit
            request.session["_dashboard_authenticated_user_id"] = request.user.id
            request.session.save()
            return redirect("dashboard:index")
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse_lazy("dashboard:index")


def dashboard_logout_view(request):
    """
    Déconnexion du dashboard UNIQUEMENT.
    
    Supprime l'indicateur `_dashboard_authenticated_user_id` de la session
    mais ne déconnecte PAS l'utilisateur Django.
    
    Résultat :
    - L'utilisateur n'a plus accès au dashboard (le middleware le redirige)
    - L'utilisateur reste connecté sur le site client s'il l'était
    - La session client (panier, etc.) est préservée
    """
    if request.user.is_authenticated:
        log_activity(
            request.user, "logout", "Dashboard",
            None, request.user.username,
            "Déconnexion du tableau de bord",
            request.META.get("REMOTE_ADDR"),
        )

    # Supprimer l'indicateur dashboard de la session
    if "_dashboard_authenticated_user_id" in request.session:
        del request.session["_dashboard_authenticated_user_id"]

    # Marquer la session comme modifiée pour forcer la sauvegarde
    request.session.modified = True
    request.session.save()

    messages.info(request, "Vous êtes déconnecté du tableau de bord.")
    return redirect("dashboard:login")


# ===== DASHBOARD HOME =====

@staff_required
def dashboard_home(request):
    orders = Order.objects.all()
    product_count = Product.objects.count()
    category_count = Category.objects.count()
    active_product_count = Product.objects.filter(is_active=True).count()
    low_stock_count = Product.objects.filter(stock__lte=5, is_active=True).count()
    featured_count = Product.objects.filter(is_featured=True, is_active=True).count()
    on_sale_count = Product.objects.filter(on_sale=True, is_active=True).count()
    banner_count = Banner.objects.filter(is_active=True).count()
    total_revenue = orders.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")

    status_breakdown = {
        "pending": orders.filter(status=Order.STATUS_PENDING).count(),
        "confirmed": orders.filter(status=Order.STATUS_CONFIRMED).count(),
        "preparing": orders.filter(status=Order.STATUS_PREPARING).count(),
        "shipped": orders.filter(status=Order.STATUS_SHIPPED).count(),
        "delivered": orders.filter(status=Order.STATUS_DELIVERED).count(),
        "cancelled": orders.filter(status=Order.STATUS_CANCELLED).count(),
    }

    recent_orders = orders.prefetch_related("items").order_by("-created_at")[:6]
    top_products = (
        Product.objects.filter(order_items__isnull=False)
        .select_related("category")
        .annotate(
            sold_quantity=Sum("order_items__quantity"),
            order_count=Count("order_items__order", distinct=True),
        )
        .order_by("-sold_quantity", "name")[:5]
    )
    low_stock_products = Product.objects.filter(
        is_active=True, stock__lte=5
    ).order_by("stock", "name")[:5]

    context = {
        "stats": {
            "order_count": orders.count(),
            "product_count": product_count,
            "category_count": category_count,
            "active_product_count": active_product_count,
            "low_stock_count": low_stock_count,
            "featured_count": featured_count,
            "on_sale_count": on_sale_count,
            "banner_count": banner_count,
            "customer_count": orders.values("email").exclude(email="").distinct().count(),
            "revenue": total_revenue,
        },
        "status_breakdown": status_breakdown,
        "recent_orders": recent_orders,
        "top_products": top_products,
        "low_stock_products": low_stock_products,
    }
    return render(request, "dashboard/index.html", context)


# ===== PRODUCTS =====

@staff_required
def product_list(request):
    q = request.GET.get("q", "").strip()
    category_filter = request.GET.get("category", "").strip()
    status_filter = request.GET.get("status", "").strip()

    products = Product.objects.select_related("category").order_by("-is_featured", "name")

    if q:
        products = products.filter(
            Q(name__icontains=q)
            | Q(short_description__icontains=q)
            | Q(sku__icontains=q)
        )
    if category_filter and category_filter.isdigit():
        products = products.filter(category_id=int(category_filter))
    if status_filter == "active":
        products = products.filter(is_active=True)
    elif status_filter == "inactive":
        products = products.filter(is_active=False)
    elif status_filter == "featured":
        products = products.filter(is_featured=True)
    elif status_filter == "on_sale":
        products = products.filter(on_sale=True)
    elif status_filter == "low_stock":
        products = products.filter(stock__lte=5, is_active=True)

    categories = Category.objects.filter(is_active=True)

    context = {
        "products": products,
        "categories": categories,
        "summary": {
            "total": Product.objects.count(),
            "active": Product.objects.filter(is_active=True).count(),
            "featured": Product.objects.filter(is_featured=True).count(),
            "on_sale": Product.objects.filter(on_sale=True).count(),
        },
    }
    return render(request, "dashboard/product_list.html", context)


@staff_required
def product_create(request):
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            log_activity(
                request.user, "create", "Produit", product.pk, product.name,
                f"Création du produit « {product.name} » (prix: {product.price} Ar, stock: {product.stock})",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Produit « {product.name} » créé avec succès.")
            return redirect("dashboard:product_list")
    else:
        form = ProductForm()

    return render(request, "dashboard/product_form.html", {"form": form})


@staff_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    old_name = product.name

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Produit", product.pk, product.name,
                f"Modification du produit « {product.name} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Produit « {product.name} » modifié.")
            return redirect("dashboard:product_list")
    else:
        form = ProductForm(instance=product)

    return render(request, "dashboard/product_form.html", {"form": form})


@staff_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        name = product.name
        pk_val = product.pk
        product.delete()
        log_activity(
            request.user, "delete", "Produit", pk_val, name,
            f"Suppression du produit « {name} »",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(request, f"Produit « {name} » supprimé.")
        return redirect("dashboard:product_list")
    return render(request, "dashboard/confirm_delete.html", {
        "object": product,
        "cancel_url": reverse_lazy("dashboard:product_list"),
    })


# ===== CATEGORIES =====

@staff_required
def category_list(request):
    q = request.GET.get("q", "").strip()

    categories = Category.objects.annotate(
        product_total=Count("products", distinct=True),
        active_product_total=Count(
            "products", filter=Q(products__is_active=True), distinct=True
        ),
    ).order_by("name")

    if q:
        categories = categories.filter(name__icontains=q)

    context = {
        "categories": categories,
        "summary": {
            "total": categories.count(),
            "active": categories.filter(is_active=True).count(),
            "inactive": categories.filter(is_active=False).count(),
        },
    }
    return render(request, "dashboard/category_list.html", context)


@staff_required
def category_create(request):
    if request.method == "POST":
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            category = form.save()
            log_activity(
                request.user, "create", "Catégorie", category.pk, category.name,
                f"Création de la catégorie « {category.name} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Catégorie « {category.name} » créée.")
            return redirect("dashboard:category_list")
    else:
        form = CategoryForm()

    return render(request, "dashboard/category_form.html", {"form": form})


@staff_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)

    if request.method == "POST":
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Catégorie", category.pk, category.name,
                f"Modification de la catégorie « {category.name} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Catégorie « {category.name} » modifiée.")
            return redirect("dashboard:category_list")
    else:
        form = CategoryForm(instance=category)

    return render(request, "dashboard/category_form.html", {"form": form})


@staff_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        name = category.name
        pk_val = category.pk
        category.delete()
        log_activity(
            request.user, "delete", "Catégorie", pk_val, name,
            f"Suppression de la catégorie « {name} »",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(request, f"Catégorie « {name} » supprimée.")
        return redirect("dashboard:category_list")
    return render(request, "dashboard/confirm_delete.html", {
        "object": category,
        "cancel_url": reverse_lazy("dashboard:category_list"),
    })


# ===== ORDERS =====

@staff_required
def order_list(request):
    q = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "").strip()

    orders = Order.objects.annotate(item_count=Count("items")).order_by("-created_at")

    if q:
        orders = orders.filter(
            Q(full_name__icontains=q)
            | Q(email__icontains=q)
            | Q(phone__icontains=q)
            | Q(id__icontains=q)
        )
    if status_filter in dict(Order.STATUS_CHOICES):
        orders = orders.filter(status=status_filter)

    context = {
        "orders": orders,
        "status_choices": Order.STATUS_CHOICES,
        "summary": {
            "total": orders.count(),
            "pending": orders.filter(status=Order.STATUS_PENDING).count(),
            "confirmed": orders.filter(status=Order.STATUS_CONFIRMED).count(),
            "delivered": orders.filter(status=Order.STATUS_DELIVERED).count(),
        },
    }
    return render(request, "dashboard/order_list.html", context)


@staff_required
def order_detail(request, pk):
    order = get_object_or_404(
        Order.objects.prefetch_related("items__product"), pk=pk
    )

    if request.method == "POST" and request.POST.get("action") == "update_status":
        new_status = request.POST.get("status")
        if new_status in dict(Order.STATUS_CHOICES):
            if order.can_transition_to(new_status):
                old_status = order.get_status_display()
                order.status = new_status
                order.save(update_fields=["status", "updated_at"])
                log_activity(
                    request.user, "order_status", "Commande", order.pk, f"Commande #{order.id}",
                    f"Changement de statut : {old_status} → {order.get_status_display()} (client: {order.full_name})",
                    request.META.get("REMOTE_ADDR"),
                )
                messages.success(
                    request,
                    f"Commande # {order.id} : statut mis à jour → {order.get_status_display()}.",
                )
            else:
                messages.error(
                    request,
                    f"Transition de statut impossible : {order.get_status_display()} → {dict(Order.STATUS_CHOICES).get(new_status, new_status)}.",
                )
        return redirect("dashboard:order_detail", pk=order.pk)

    # Filtrer les statuts disponibles pour la transition
    valid_statuses = [(s, label) for s, label in Order.STATUS_CHOICES if order.can_transition_to(s)]

    context = {
        "order": order,
        "status_choices": Order.STATUS_CHOICES,
        "valid_statuses": valid_statuses,
    }
    return render(request, "dashboard/order_detail.html", context)


# ===== BANNERS =====

@staff_required
def banner_list(request):
    banners = Banner.objects.all().order_by("position", "order", "-created_at")
    context = {
        "banners": banners,
        "summary": {
            "total": banners.count(),
            "active": banners.filter(is_active=True).count(),
            "hero": banners.filter(position="hero").count(),
        },
    }
    return render(request, "dashboard/banner_list.html", context)


@staff_required
def banner_create(request):
    if request.method == "POST":
        form = BannerForm(request.POST, request.FILES)
        if form.is_valid():
            banner = form.save()
            log_activity(
                request.user, "create", "Bannière", banner.pk, banner.title or "Sans titre",
                f"Création de la bannière « {banner.title or 'Sans titre'} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Bannière « {banner.title or 'Sans titre'} » créée.")
            return redirect("dashboard:banner_list")
    else:
        form = BannerForm()

    return render(request, "dashboard/banner_form.html", {"form": form, "title": "Nouvelle bannière"})


@staff_required
def banner_edit(request, pk):
    banner = get_object_or_404(Banner, pk=pk)
    if request.method == "POST":
        form = BannerForm(request.POST, request.FILES, instance=banner)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Bannière", banner.pk, banner.title or "Sans titre",
                f"Modification de la bannière « {banner.title or 'Sans titre'} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Bannière « {banner.title or 'Sans titre'} » modifiée.")
            return redirect("dashboard:banner_list")
    else:
        form = BannerForm(instance=banner)

    return render(request, "dashboard/banner_form.html", {"form": form, "title": "Modifier la bannière"})


@staff_required
def banner_delete(request, pk):
    banner = get_object_or_404(Banner, pk=pk)
    if request.method == "POST":
        title = banner.title or "Sans titre"
        pk_val = banner.pk
        banner.delete()
        log_activity(
            request.user, "delete", "Bannière", pk_val, title,
            f"Suppression de la bannière « {title} »",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(request, f"Bannière « {title} » supprimée.")
        return redirect("dashboard:banner_list")
    return render(request, "dashboard/confirm_delete.html", {
        "object": banner,
        "cancel_url": reverse_lazy("dashboard:banner_list"),
    })


# ===== SITE SETTINGS =====

@staff_required
def site_settings(request):
    config = SiteConfiguration.load()

    if request.method == "POST":
        form = SiteConfigurationForm(request.POST, request.FILES, instance=config)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Configuration du site",
                config.pk, config.site_name,
                "Modification des paramètres du site",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, "Configuration du site mise à jour.")
            return redirect("dashboard:site_settings")
    else:
        form = SiteConfigurationForm(instance=config)

    return render(request, "dashboard/site_settings.html", {"form": form})


# ===== USERS =====

@staff_required
def user_list(request):
    q = request.GET.get("q", "").strip()
    role_filter = request.GET.get("role", "").strip()

    users = User.objects.select_related("profile").order_by("-date_joined")

    if q:
        users = users.filter(
            Q(username__icontains=q)
            | Q(email__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
        )
    if role_filter == "staff":
        users = users.filter(is_staff=True)
    elif role_filter == "active":
        users = users.filter(is_active=True)
    elif role_filter == "inactive":
        users = users.filter(is_active=False)

    context = {
        "users": users,
        "summary": {
            "total": User.objects.count(),
            "staff": User.objects.filter(is_staff=True).count(),
            "active": User.objects.filter(is_active=True).count(),
        },
    }
    return render(request, "dashboard/user_list.html", context)


@staff_required
def user_create(request):
    if request.method == "POST":
        form = UserDashboardForm(request.POST)
        if form.is_valid():
            user = form.save()
            log_activity(
                request.user, "create", "Utilisateur", user.pk, user.username,
                f"Création de l'utilisateur « {user.username} » (rôle: {'Admin' if user.is_superuser else 'Client'})",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Utilisateur « {user.username} » créé avec succès.")
            return redirect("dashboard:user_list")
    else:
        form = UserDashboardForm()

    return render(request, "dashboard/user_form.html", {"form": form, "title": "Nouvel utilisateur"})


@staff_required
def user_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        form = UserDashboardForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Utilisateur", user.pk, user.username,
                f"Modification de l'utilisateur « {user.username} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Utilisateur « {user.username} » modifié.")
            return redirect("dashboard:user_list")
    else:
        form = UserDashboardForm(instance=user)

    return render(request, "dashboard/user_form.html", {"form": form, "title": f"Modifier : {user.username}"})


@staff_required
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        username = user.username
        pk_val = user.pk
        user.delete()
        log_activity(
            request.user, "delete", "Utilisateur", pk_val, username,
            f"Suppression de l'utilisateur « {username} »",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(request, f"Utilisateur « {username} » supprimé.")
        return redirect("dashboard:user_list")
    return render(request, "dashboard/confirm_delete.html", {
        "object": user,
        "cancel_url": reverse_lazy("dashboard:user_list"),
    })


# ===== COUPONS =====

@staff_required
def coupon_list(request):
    q = request.GET.get("q", "").strip()
    coupons = Coupon.objects.all().order_by("-created_at")
    if q:
        coupons = coupons.filter(code__icontains=q)
    context = {
        "coupons": coupons,
        "summary": {
            "total": Coupon.objects.count(),
            "active": Coupon.objects.filter(is_active=True).count(),
        },
    }
    return render(request, "dashboard/coupon_list.html", context)


@staff_required
def coupon_create(request):
    if request.method == "POST":
        form = CouponForm(request.POST)
        if form.is_valid():
            coupon = form.save()
            log_activity(
                request.user, "create", "Coupon", coupon.pk, coupon.code,
                f"Création du coupon « {coupon.code} » (réduction: {coupon.discount_value}{'%' if coupon.discount_type == 'percentage' else ' Ar'})",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Coupon « {coupon.code} » créé.")
            return redirect("dashboard:coupon_list")
    else:
        form = CouponForm()

    return render(request, "dashboard/coupon_form.html", {"form": form, "title": "Nouveau coupon"})


@staff_required
def coupon_edit(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    if request.method == "POST":
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Coupon", coupon.pk, coupon.code,
                f"Modification du coupon « {coupon.code} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Coupon « {coupon.code} » modifié.")
            return redirect("dashboard:coupon_list")
    else:
        form = CouponForm(instance=coupon)

    return render(request, "dashboard/coupon_form.html", {"form": form, "title": f"Modifier : {coupon.code}"})


@staff_required
def coupon_delete(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    if request.method == "POST":
        code = coupon.code
        pk_val = coupon.pk
        coupon.delete()
        log_activity(
            request.user, "delete", "Coupon", pk_val, code,
            f"Suppression du coupon « {code} »",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(request, f"Coupon « {code} » supprimé.")
        return redirect("dashboard:coupon_list")
    return render(request, "dashboard/confirm_delete.html", {
        "object": coupon,
        "cancel_url": reverse_lazy("dashboard:coupon_list"),
    })


# ===== SHIPPING ZONES =====

@staff_required
def shipping_list(request):
    zones = ShippingZone.objects.all().order_by("order", "name")
    context = {
        "shipping_zones": zones,
        "summary": {
            "total": zones.count(),
            "active": zones.filter(is_active=True).count(),
        },
    }
    return render(request, "dashboard/shipping_list.html", context)


@staff_required
def shipping_create(request):
    if request.method == "POST":
        form = ShippingZoneForm(request.POST)
        if form.is_valid():
            zone = form.save()
            log_activity(
                request.user, "create", "Zone de livraison", zone.pk, zone.name,
                f"Création de la zone de livraison « {zone.name} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Zone de livraison « {zone.name} » créée.")
            return redirect("dashboard:shipping_list")
    else:
        form = ShippingZoneForm()

    return render(request, "dashboard/shipping_form.html", {"form": form, "title": "Nouvelle zone de livraison"})


@staff_required
def shipping_edit(request, pk):
    zone = get_object_or_404(ShippingZone, pk=pk)
    if request.method == "POST":
        form = ShippingZoneForm(request.POST, instance=zone)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Zone de livraison", zone.pk, zone.name,
                f"Modification de la zone de livraison « {zone.name} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Zone de livraison « {zone.name} » modifiée.")
            return redirect("dashboard:shipping_list")
    else:
        form = ShippingZoneForm(instance=zone)

    return render(request, "dashboard/shipping_form.html", {"form": form, "title": f"Modifier : {zone.name}"})


@staff_required
def shipping_delete(request, pk):
    zone = get_object_or_404(ShippingZone, pk=pk)
    if request.method == "POST":
        name = zone.name
        pk_val = zone.pk
        zone.delete()
        log_activity(
            request.user, "delete", "Zone de livraison", pk_val, name,
            f"Suppression de la zone de livraison « {name} »",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(request, f"Zone de livraison « {name} » supprimée.")
        return redirect("dashboard:shipping_list")
    return render(request, "dashboard/confirm_delete.html", {
        "object": zone,
        "cancel_url": reverse_lazy("dashboard:shipping_list"),
    })


# ===== TAX RATES =====

@staff_required
def tax_list(request):
    taxes = TaxRate.objects.all().order_by("-rate")
    context = {
        "taxes": taxes,
        "summary": {
            "total": taxes.count(),
            "active": taxes.filter(is_active=True).count(),
        },
    }
    return render(request, "dashboard/tax_list.html", context)


@staff_required
def tax_create(request):
    if request.method == "POST":
        form = TaxRateForm(request.POST)
        if form.is_valid():
            tax = form.save()
            log_activity(
                request.user, "create", "Taxe", tax.pk, tax.name,
                f"Création de la taxe « {tax.name} » ({tax.rate}%)",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Taxe « {tax.name} » créée.")
            return redirect("dashboard:tax_list")
    else:
        form = TaxRateForm()

    return render(request, "dashboard/tax_form.html", {"form": form, "title": "Nouveau taux de taxe"})


@staff_required
def tax_edit(request, pk):
    tax = get_object_or_404(TaxRate, pk=pk)
    if request.method == "POST":
        form = TaxRateForm(request.POST, instance=tax)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Taxe", tax.pk, tax.name,
                f"Modification de la taxe « {tax.name} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Taxe « {tax.name} » modifiée.")
            return redirect("dashboard:tax_list")
    else:
        form = TaxRateForm(instance=tax)

    return render(request, "dashboard/tax_form.html", {"form": form, "title": f"Modifier : {tax.name}"})


@staff_required
def tax_delete(request, pk):
    tax = get_object_or_404(TaxRate, pk=pk)
    if request.method == "POST":
        name = tax.name
        pk_val = tax.pk
        tax.delete()
        log_activity(
            request.user, "delete", "Taxe", pk_val, name,
            f"Suppression de la taxe « {name} »",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(request, f"Taxe « {name} » supprimée.")
        return redirect("dashboard:tax_list")
    return render(request, "dashboard/confirm_delete.html", {
        "object": tax,
        "cancel_url": reverse_lazy("dashboard:tax_list"),
    })


# ===== ACTIVITY LOG =====

@staff_required
def activity_log(request):
    logs = ActivityLog.objects.select_related("user").order_by("-created_at")
    action_filter = request.GET.get("action", "").strip()
    search_query = request.GET.get("q", "").strip()
    export_excel = request.GET.get("export_excel", "").strip()
    
    if action_filter in dict(ActivityLog.ACTION_CHOICES):
        logs = logs.filter(action=action_filter)
    if search_query:
        logs = logs.filter(
            Q(user__username__icontains=search_query)
            | Q(object_repr__icontains=search_query)
            | Q(details__icontains=search_query)
            | Q(model_name__icontains=search_query)
        )
    
    # Export Excel
    if export_excel == "1":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Journal d'activité"
        
        # En-têtes
        headers = ["Date", "Utilisateur", "Action", "Modèle", "Objet", "Détails", "Adresse IP"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row, log in enumerate(logs, 2):
            ws.cell(row=row, column=1, value=log.created_at.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row, column=2, value=log.user.username if log.user else "Système")
            ws.cell(row=row, column=3, value=log.get_action_display())
            ws.cell(row=row, column=4, value=log.model_name or "")
            ws.cell(row=row, column=5, value=log.object_repr or "")
            ws.cell(row=row, column=6, value=log.details or "")
            ws.cell(row=row, column=7, value=log.ip_address or "")
        
        # Ajuster la largeur des colonnes
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 25
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="journal_activite.xlsx"'
        wb.save(response)
        return response
    
    # Pagination
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(logs, 50)
    page = request.GET.get("page", 1)
    try:
        logs_page = paginator.page(page)
    except PageNotAnInteger:
        logs_page = paginator.page(1)
    except EmptyPage:
        logs_page = paginator.page(paginator.num_pages)
    
    context = {
        "logs": logs_page,
        "action_choices": ActivityLog.ACTION_CHOICES,
        "paginator": paginator,
    }
    return render(request, "dashboard/activity_log.html", context)


@staff_required
def activity_log_detail(request, pk):
    """Affiche les détails complets d'une entrée du journal."""
    log_entry = get_object_or_404(ActivityLog, pk=pk)
    return render(request, "dashboard/activity_log_detail.html", {"log": log_entry})


@require_POST
@staff_required
def activity_log_delete(request, pk):
    """Supprime une entrée du journal d'activité."""
    log_entry = get_object_or_404(ActivityLog, pk=pk)
    log_entry.delete()
    messages.success(request, "Entrée du journal supprimée.")
    return redirect("dashboard:activity_log")


# ===== PAYMENT METHODS =====

@staff_required
def payment_list(request):
    methods = PaymentMethod.objects.all().order_by("method")
    context = {
        "methods": methods,
        "summary": {
            "total": methods.count(),
            "active": methods.filter(is_active=True).count(),
        },
    }
    return render(request, "dashboard/payment_list.html", context)


@staff_required
def payment_create(request):
    if request.method == "POST":
        form = PaymentMethodForm(request.POST)
        if form.is_valid():
            pm = form.save()
            log_activity(
                request.user, "create", "Mode de paiement", pm.pk, pm.get_method_display(),
                f"Création du mode de paiement « {pm.get_method_display()} » ({pm.phone_number})",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Mode de paiement « {pm.get_method_display()} » créé.")
            return redirect("dashboard:payment_list")
    else:
        form = PaymentMethodForm()

    return render(request, "dashboard/payment_form.html", {"form": form, "title": "Nouveau mode de paiement"})


@staff_required
def payment_edit(request, pk):
    pm = get_object_or_404(PaymentMethod, pk=pk)
    if request.method == "POST":
        form = PaymentMethodForm(request.POST, instance=pm)
        if form.is_valid():
            form.save()
            log_activity(
                request.user, "update", "Mode de paiement", pm.pk, pm.get_method_display(),
                f"Modification du mode de paiement « {pm.get_method_display()} »",
                request.META.get("REMOTE_ADDR"),
            )
            messages.success(request, f"Mode de paiement « {pm.get_method_display()} » modifié.")
            return redirect("dashboard:payment_list")
    else:
        form = PaymentMethodForm(instance=pm)

    return render(request, "dashboard/payment_form.html", {"form": form, "title": f"Modifier : {pm.get_method_display()}"})


@staff_required
def payment_delete(request, pk):
    pm = get_object_or_404(PaymentMethod, pk=pk)
    if request.method == "POST":
        method_name = pm.get_method_display()
        pk_val = pm.pk
        pm.delete()
        log_activity(
            request.user, "delete", "Mode de paiement", pk_val, method_name,
            f"Suppression du mode de paiement « {method_name} »",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(request, f"Mode de paiement « {method_name} » supprimé.")
        return redirect("dashboard:payment_list")
    return render(request, "dashboard/confirm_delete.html", {
        "object": pm,
        "cancel_url": reverse_lazy("dashboard:payment_list"),
    })


# ===== PASSWORD RESET REQUESTS =====

@staff_required
def password_reset_request_list(request):
    """Liste des demandes de réinitialisation de mot de passe."""
    requests = PasswordResetRequest.objects.select_related("user", "resolved_by").all()
    pending = requests.filter(is_resolved=False).count()
    resolved = requests.filter(is_resolved=True).count()
    context = {
        "requests": requests,
        "summary": {
            "total": requests.count(),
            "pending": pending,
            "resolved": resolved,
        },
    }
    return render(request, "dashboard/password_reset_list.html", context)


@staff_required
def password_reset_send(request, pk):
    """L'admin génère un nouveau mot de passe et l'envoie par email."""
    from accounts.views import generate_random_password
    from django.core.mail import send_mail
    from django.utils import timezone

    reset_req = get_object_or_404(PasswordResetRequest, pk=pk, is_resolved=False)

    if not reset_req.user:
        messages.error(request, "Aucun utilisateur associé à cette demande.")
        return redirect("dashboard:password_reset_list")

    # Générer un nouveau mot de passe
    new_password = generate_random_password()
    reset_req.user.set_password(new_password)
    reset_req.user.save()

    # Envoyer l'email
    subject = f"Réinitialisation de votre mot de passe - {settings.SITE_NAME}"
    message = (
        f"Bonjour {reset_req.user.username},\n\n"
        f"Votre demande de réinitialisation de mot de passe a été traitée.\n\n"
        f"Voici votre nouveau mot de passe : {new_password}\n\n"
        f"Vous pouvez vous connecter et le modifier dans votre profil.\n\n"
        f"Cordialement,\n"
        f"L'équipe {settings.SITE_NAME}"
    )

    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [reset_req.email],
            fail_silently=False,
        )
        # Marquer comme résolue
        reset_req.is_resolved = True
        reset_req.resolved_at = timezone.now()
        reset_req.resolved_by = request.user
        reset_req.new_password = new_password
        reset_req.save()
        log_activity(
            request.user, "update", "Réinitialisation mot de passe",
            reset_req.pk, reset_req.email,
            f"Réinitialisation du mot de passe pour {reset_req.email}",
            request.META.get("REMOTE_ADDR"),
        )
        messages.success(
            request,
            f"Nouveau mot de passe envoyé à {reset_req.email}. "
            f"Mot de passe généré : {new_password}",
        )
    except Exception as e:
        messages.error(
            request,
            f"Erreur lors de l'envoi de l'email : {e}",
        )

    return redirect("dashboard:password_reset_list")


# ===== BACKUP =====

@staff_required
def database_backup(request):
    """Télécharge une copie de sauvegarde de la base de données."""
    db_path = settings.DATABASES["default"]["NAME"]
    db_file = Path(db_path)

    if not db_file.exists():
        messages.error(request, "Fichier de base de données introuvable.")
        return redirect("dashboard:index")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"backup_{timestamp}.sqlite3"

    log_activity(
        request.user, "other", "Sauvegarde",
        None, "Base de données",
        f"Téléchargement de la sauvegarde de la base de données",
        request.META.get("REMOTE_ADDR"),
    )

    response = FileResponse(
        open(db_file, "rb"),
        content_type="application/octet-stream",
        as_attachment=True,
        filename=filename,
    )
    return response